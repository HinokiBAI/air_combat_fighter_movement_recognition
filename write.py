import pandas as pd
import os
from openpyxl import Workbook


class Climb:
    def __init__(self):
        self.times = []
        self.id = None
        self.action = '爬升'

    def work(self, file):
        self.times = []
        data = pd.read_csv(file, usecols=['Unix time', 'Id','Roll', 'Pitch', 'Yaw', 'Longitude', 'Latitude','Altitude'])
        data.set_index('Unix time', inplace=True, drop=False)
        self.id = data['Id'].to_list()[0]
        # 找到 Roll 在 -20 与 -70 之间的时间
        for i in range(len(data)):
            if data.iloc[i]['Pitch'] > 7.5 and data.iloc[i]['Pitch'] < 80:
                self.times.append(data.iloc[i]['Unix time'])

class Swoop:
    def __init__(self):
        self.times = []
        self.id = None
        self.action = '俯冲'

    def work(self, file):
        self.times = []
        data = pd.read_csv(file, usecols=['Unix time', 'Id','Roll', 'Pitch', 'Yaw', 'Longitude', 'Latitude','Altitude'])
        data.set_index('Unix time', inplace=True, drop=False)
        self.id = data['Id'].to_list()[0]
        # 找到 Roll 在 -20 与 -70 之间的时间
        for i in range(len(data)):
            if data.iloc[i]['Pitch'] > -80 and data.iloc[i]['Pitch'] < -7.5:
                self.times.append(data.iloc[i]['Unix time'])

class Roll:
    def __init__(self):
        self.times = []
        self.id = None
        self.action = '横滚'

    def work(self, file):
        self.times = []
        data = pd.read_csv(file, usecols=['Unix time', 'Id','Roll', 'Pitch', 'Yaw', 'Longitude', 'Latitude','Altitude'])
        data.set_index('Unix time', inplace=True, drop=False)
        self.id = data['Id'].to_list()[0]
        times_ = []
        threshold = 110
        for i in range(1, len(data) - 2):
            # 计算两两横滚角的差值
            temp = data.iloc[i+1]['Roll'] - data.iloc[i]['Roll']
            if abs(temp) > threshold:
                times_.append([data.iloc[i-1]['Unix time'], data.iloc[i]['Unix time'], data.iloc[i+1]['Unix time'], data.iloc[i+2]['Unix time']])
        
        for time_ in times_:
            self.times.extend(time_)

class Turn:
    def __init__(self):
        self.times = []
        self.id = None
        self.action = '转弯'

    def work(self, file):
        self.times = []
        data = pd.read_csv(file, usecols=['Unix time', 'Id','Roll', 'Pitch', 'Yaw', 'Longitude', 'Latitude','Altitude'])
        data.set_index('Unix time', inplace=True, drop=False)
        self.id = data['Id'].to_list()[0]
        times_ = []
        times = []
        # 计算偏航角变化率
        data['Yaw_rate'] = data['Yaw'].diff()
        data['Yaw_rate'] = data['Yaw_rate'].fillna(0)

        # 找到 Roll 在 -20 与 -70 之间的时间
        for i in range(len(data)):
            if (-80 < data.iloc[i]['Roll'] < -45 or 45 < data.iloc[i]['Roll'] < 80) and abs(data.iloc[i]['Yaw_rate']) > 2.5:
                times.append(data.iloc[i]['Unix time'])


        # 遍历times, 转换为times_
        times_ = []
        temp = []
        for i in range(len(times)):
            temp.append(times[i])
            if times[i] + 1 not in times:
                times_.append(temp)
                temp = []

        # 求俯仰角累积变化量
        pitch_acc = []
        for i in range(len(times_)):
            temp = 0
            for j in range(len(times_[i]) - 1):
                temp += abs(data.loc[times_[i][j+1]]['Yaw'] - data.loc[times_[i][j]]['Yaw'])
            pitch_acc.append(temp)

        flag = []
        for i in range(len(pitch_acc)):
            if pitch_acc[i] > 40:
                flag.append(i)

        for i in flag:
            self.times.extend(times_[i])

class Somersault:
    def __init__(self):
        self.times = []
        self.id = None
        self.action = '筋斗'

    def work(self, file):
        self.times = []
        data = pd.read_csv(file, usecols=['Unix time', 'Id','Roll', 'Pitch', 'Yaw', 'Longitude', 'Latitude','Altitude'])
        data.set_index('Unix time', inplace=True, drop=False)
        self.id = data['Id'].to_list()[0]
        times_ = []
        times = []

        data['Pitch_rate'] = data['Pitch'].diff()
        data['Pitch_rate'] = data['Pitch_rate'].fillna(0)
        
        times = []

        # 找到 Roll 在 -20 与 -70 之间的时间
        for i in range(len(data)):
            if (data.iloc[i]['Pitch_rate'] > 6 and data.iloc[i]['Pitch_rate'] < 10) or(data.iloc[i]['Pitch_rate'] > -10 and data.iloc[i]['Pitch_rate'] < -6):
                times.append(data.iloc[i]['Unix time'])


        # 遍历times, 转换为times_
        times_ = []
        temp = []
        for i in range(len(times)):
            temp.append(times[i])
            if times[i] + 1 not in times:
                times_.append(temp)
                temp = []

        # 拼接times_
        # 阈值
        threshold = 10
        for i in range(len(times_)-1):
            for j in range(1, threshold):
                if times_[i][-1] + j in times_[i+1]:
                    for k in range(j):
                        times_[i].append(times_[i][-1] + k + 1)

                    times_[i].extend(times_[i+1])
                    times_[i+1] = times_[i]
                    break

        # 求俯仰角累积变化量
        pitch_acc = []
        for i in range(len(times_)):
            temp = 0
            for j in range(len(times_[i]) - 1):
                try:
                    temp += abs(data.loc[times_[i][j+1]]['Pitch'] - data.loc[times_[i][j]]['Pitch'])
                except:
                    continue
            pitch_acc.append(temp)

        flag = []
        for i in range(len(pitch_acc)):
            if pitch_acc[i] > 200:
                flag.append(i)

        for i in flag:
            self.times.extend(times_[i])

class Convolve:
    def __init__(self) -> None:
        self.times = []
        self.id = None
        self.action = '盘旋'

    def work(self, file):
        self.times = []
        data = pd.read_csv(file, usecols=['Unix time', 'Id','Roll', 'Pitch', 'Yaw', 'Longitude', 'Latitude','Altitude'])
        data.set_index('Unix time', inplace=True, drop=False)
        self.id = data['Id'].to_list()[0]
        times_ = []
        times = []
        
        # 计算偏航角变化率
        data['Yaw_rate'] = data['Yaw'].diff()
        data['Yaw_rate'] = data['Yaw_rate'].fillna(0)

        # 找到 Roll 在 -20 与 -70 之间的时间
        for i in range(len(data)):
            if (-80 < data.iloc[i]['Roll'] < -20 or 20 < data.iloc[i]['Roll'] < 80) and data.iloc[i]['Yaw_rate'] > 2 :
                times.append(data.iloc[i]['Unix time'])


        # 遍历times, 转换为times_
        times_ = []
        temp = []
        for i in range(len(times)):
            temp.append(times[i])
            if times[i] + 1 not in times:
                times_.append(temp)
                temp = []

        # 求俯仰角累积变化量
        pitch_acc = []
        for i in range(len(times_)):
            temp = 0
            for j in range(len(times_[i]) - 1):
                if data.loc[times_[i][j+1]]['Yaw'] * data.loc[times_[i][j]]['Yaw'] < 0:
                    temp += abs(data.loc[times_[i][j+1]]['Yaw'] + data.loc[times_[i][j]]['Yaw'])
                else:
                    temp += abs(data.loc[times_[i][j+1]]['Yaw'] - data.loc[times_[i][j]]['Yaw'])
            pitch_acc.append(temp)



        altitude_change = []
        for i in range(len(times_)):
            temp = abs(data.loc[times_[i][-1]]['Altitude'] - data.loc[times_[i][0]]['Altitude'])
            altitude_change.append(temp)

        flag = []
        for i in range(len(pitch_acc)):
            if pitch_acc[i] >200 and altitude_change[i] <= 500:
                flag.append(i)

        for i in flag:
            self.times.extend(times_[i])

class File:
    def __init__(self, path):
        self.path = path
        self.climb = Climb()
        self.swoop = Swoop()
        self.roll = Roll()
        self.turn = Turn()
        self.somersault = Somersault()
        self.convolve = Convolve()
        self.book = Workbook()
        self.sheet = self.book.active
        self.col = 1
        self.init()

    def init(self):
        # 初始化sheet
        self.sheet.cell(1, 1).value = 'Time/ID'
        for i in range(1, 3000):
            self.sheet.cell(i + 1, 1).value = i
        self.save()

    def get_Plane_file(self):
        mydir = os.path.join(self.path, 'Plane')
        # 获取 Plane 目录下的所有文件
        files = os.listdir(mydir)
        for file in files:
            self.col += 1
            self.climb.work(os.path.join(mydir, file))
            self.write_file(self.climb.id, self.climb.times, self.climb.action)
            self.swoop.work(os.path.join(mydir, file))
            self.write_file(self.swoop.id, self.swoop.times, self.swoop.action)
            self.roll.work(os.path.join(mydir, file))
            self.write_file(self.roll.id, self.roll.times, self.roll.action)
            self.turn.work(os.path.join(mydir, file))
            self.write_file(self.turn.id, self.turn.times, self.turn.action)
            self.somersault.work(os.path.join(mydir, file))
            self.write_file(self.somersault.id, self.somersault.times, self.somersault.action)
            self.convolve.work(os.path.join(mydir, file))
            self.write_file(self.convolve.id, self.convolve.times, self.convolve.action)


    def get_Weapon_file(self):
        mydir = os.path.join(self.path, 'Weapon')
        # 获取 Weapon 目录下的所有文件
        files = os.listdir(mydir)
        for file in files:
            self.col += 1
            self.climb.work(os.path.join(mydir, file))
            self.write_file(self.climb.id, self.climb.times, self.climb.action)
            self.swoop.work(os.path.join(mydir, file))
            self.write_file(self.swoop.id, self.swoop.times, self.swoop.action)
            self.roll.work(os.path.join(mydir, file))
            self.write_file(self.roll.id, self.roll.times, self.roll.action)
            self.turn.work(os.path.join(mydir, file))
            self.write_file(self.turn.id, self.turn.times, self.turn.action)
            self.somersault.work(os.path.join(mydir, file))
            self.write_file(self.somersault.id, self.somersault.times, self.somersault.action)
            self.convolve.work(os.path.join(mydir, file))
            self.write_file(self.convolve.id, self.convolve.times, self.convolve.action)

    def write_file(self, id, times, action):
        # 获取sheet列数
        self.sheet.cell(1, self.col).value = id
        for time in times:
            self.sheet.cell(time + 1, self.col).value = action
        print(id ,action, '保存成功！')

    def save(self):
        self.book.save(os.path.join(self.path, os.path.basename(self.path) + '.xlsx'))

    def work(self):
        self.get_Plane_file()
        self.get_Weapon_file()
        self.save()

        
if __name__ == '__main__':
    file = File(r'data\51stKIAP_vs_107th_Round_1')
    file.work()


        